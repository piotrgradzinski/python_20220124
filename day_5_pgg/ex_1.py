"""
Napisz program, który wczyta pracowników z pliku employees.xlsx i policzy:
- sumę wynagrodzeń
- średnia wynagrodzeń

Dane wyświetl w konsoli i zapisz do pliku employees_report.json.
"""

from openpyxl import load_workbook
import json

wb = load_workbook('day_5_pgg/employees.xlsx')
ws = wb.active

# salaries = []
# for row in ws.iter_rows(min_row=2):
    # salaries.append(row[2].value)

salaries = [row[2].value for row in ws.iter_rows(min_row=2)]

print(salaries)

salaries_sum = sum(salaries)
salaries_mean = salaries_sum / len(salaries)

print(f'Salaries sum: {salaries_sum:.2f}')
print(f'Salaries mean: {salaries_mean:.2f}')

with open('day_5_pgg/employees_report.json', 'w') as file_handle:
    json_data = {
        'salaries_sum': salaries_sum,
        'salaries_mean': salaries_mean,
    }
    json.dump(json_data, file_handle, indent=4)
