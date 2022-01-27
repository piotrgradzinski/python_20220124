"""
Obsluga plikow w pythonie
- otwieramy plik - oddaje nam uchwyt do pliku
    - jezeli nie ma pliku, nie mozna go znalezc to dostajemy FileNotFoundError
    - przy otwarciu pliku, moze podac tryb w jakim go otwieramy, standardowo to tryb 'r', czyli do odczytu
    - https://docs.python.org/3/library/functions.html#open
    - https://stackoverflow.com/a/16212401/779084 
    - r - czytanie (domyślny), FileNotFoundError jak nie ma pliku, wskaźnik na początku
    - w - pisanie, nadpisuje istniejący plik, tworzy nowy plik jak nie ma
    - a - dodawanie, wskaźnik na końcu pliku, tworzy nowy plik jak nie ma
    - r+ - czytanie/pisanie, wskaźnik na początku, FileNotFoundError jak nie ma pliku, nie usuwa calej zawartosci pliku
    - w+ - czytanie/pisanie, czysci zawartosc istniejacego pliku juz przy jego otwarciu, tworzy nowy plik jak nie ma
    - a+ - czytanie/dodawanie, wskaźnik na końcu pliku, tworzy nowy plik jak nie ma
    - rb, rb+, wb, wb+, ab, ab+ - analogicznie, ale w trybie binarnym

- wykonujemy operacje (zapis, odczyt)
- zamykamy plik - moze byc tak, ze jak nie zamkniecie pliku, to pozniej inny program nie bedzie w stanie z niego skorzystac
"""


file_handle = open('day_4_pgg/test.txt')
content = file_handle.read()
print(content)
file_handle.close()

print('-' * 60)

file_handle = open('day_4_pgg/test.txt', 'w+')  # tryb do zapisu
file_handle.write('Ala ma kota\nKot ma kompilator')
file_handle.close()

print('-' * 60)

"""
zeby nie trzeba bylo pamietac o zamykaniu pliku mozemy wykorzystac compound statement
takie samo rozwiazanie mozemy zastowac nie tylko do plikwo, ale takze do innych zasobow
np. requesty do API, zasoby sieciowe, itp.
"""

with open('day_4_pgg/test.txt', 'r+', encoding='utf-8') as file_handle:
    content = file_handle.read()
    print(content)
    print(file_handle.closed)  # False
    print(file_handle.encoding)  # UTF-8
    print(file_handle.mode)  # r+
    print(file_handle.name)  # day_4_pgg/test.txt
    print(file_handle.writable())  # True
    print(file_handle.readable())  # True

print(file_handle.closed)  # True

print('Jestem poza with')

print('-' * 60)

with open('day_4_pgg/test.txt', 'r+', encoding='utf-8') as file_handle:
    # jak mozemy odczytywac dane z pliku
    content = file_handle.read()  # trzeba uwazac na wczytywanie duzych plikow
    print(content)

    print('---')

    file_handle.seek(0)  # 1 - czytamy od drugiego znaku z pliku
    content = file_handle.read()  # trzeba uwazac na wczytywanie duzych plikow
    print(content)

    print('---')

    file_handle.seek(0)
    content = file_handle.read(10)
    print(content)
    content = file_handle.read(10)
    print(content)

    print('---')

    file_handle.seek(0)
    content = file_handle.read(10)
    print(content)

    print('---')

    file_handle.seek(0)
    content = file_handle.readline()  # "Ala ma kota\n" - znak \n jest rowniez dolaczany
    print(content)
    content = file_handle.readline()
    print(content)

    print('---')

    file_handle.seek(0)
    content = file_handle.readline().rstrip('\n')  # usuniecie znaku nowej linii z konca wczytanej linii
    print(content)
    content = file_handle.readline()
    print(content)

print('---')

# obrabianie duzych plikow
with open('day_4_pgg/test.txt', 'r+', encoding='utf-8') as file_handle:
    for line in file_handle:
        print(line)


print('---')

with open('day_4_pgg/test.txt', 'r+', encoding='utf-8') as file_handle:
    list_of_lines = list(file_handle)
    print(list_of_lines)

print('---')

"""
CSV
"""

import csv

# zapis do CSV
csv_data = [
    ['first_name', 'last_name', 'salary'],
    ['Piotr', 'GG', 1200],
    ['Ala', 'Nowak', 25000],
    ['Tomasz', 'Kowalski', 5200],
    ['Jan', 'Nowak', 8000],
    ['Krystyna', 'Tomaszewska', 10000],
]

with open('day_4_pgg/employees.csv', 'w') as file_handle:
    writer = csv.writer(file_handle, dialect='excel', delimiter=';')

    for row in csv_data:
        writer.writerow(row)

# odczyt z CSV
with open('day_4_pgg/employees.csv') as file_handle:
    data_from_csv = csv.DictReader(file_handle, delimiter=';')

    for row in data_from_csv:
        print(row)


"""
Excel
openpyxl
https://pypi.org/project/openpyxl/
pip install openpyxl
"""

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

xls_headers = ['first_name', 'last_name', 'salary']
xls_data = [
    ['Piotr', 'GG', 1200],
    ['Ala', 'Nowak', 25000],
    ['Tomasz', 'Kowalski', 5200],
    ['Jan', 'Nowak', 8000],
    ['Krystyna', 'Tomaszewska', 10000],
]

ws.append(xls_headers)

for row in xls_data:
    ws.append(row)

wb.save('day_4_pgg/employees.xlsx')

print('---')

# wczytywanie XLSX
from openpyxl import load_workbook

wb = load_workbook('day_4_pgg/employees2.xlsx')

for row in wb['Sheet2'].iter_rows(min_row=2):
    print(row[0].value, row[1].value, row[2].value)

print('---')

print(wb.sheetnames)  # ['Sheet', 'Sheet2']

print(wb['Sheet']['A3'].value)

print('---')

for col in wb['Sheet']['A:B']:
    for cell in col:
        print(cell.value)


print('-' * 60)

"""
Formaty przechowywania danych
- JSON - JavaScript Object Notation, https://pl.wikipedia.org/wiki/JSON
- XML
- YAML
"""

# JSON
import json

python_data = {
    'first_name': 'Piotr',
    'last_name': 'GG',
    'favourite_numbers': [1, 2, 3, 4, 5, 6],
    'is_trainer': True,
    'favourite_brands': ('PAYBACK', "ALX"),
    'additional_info': None,
    # 'test': {1, 2, 3, 4, 5}  # zbioru nie bedzie w stanie zapisac do JSONa
}

with open('day_4_pgg/data.json', 'w') as file_handle:
    print(json.dump(python_data, file_handle, indent=4))

from pprint import pprint

with open('day_4_pgg/data.json') as file_handle:
    restored_python_data = json.load(file_handle)
    pprint(restored_python_data)


print('------')

print(json.dumps(python_data, indent=4))


"""
Zadawanie zapytan w strukturze JSON
jmespath - https://pypi.org/project/jmespath/
jsonpath-ng - https://pypi.org/project/jsonpath-ng/
jq - https://pypi.org/project/jq/ , https://stedolan.github.io/jq/
"""

"""
YAML - https://pl.wikipedia.org/wiki/YAML

pip install PyYAML
"""

import yaml

with open('day_4_pgg/data.yaml', 'w') as file_handle:
    yaml.dump(python_data, file_handle)